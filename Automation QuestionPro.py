"""
    Truong Pham - Project Leader
    Email: phamduytruong159@gmail.com
    
    Divine Mbamara - Associate
    Email: divinembamara@gmail.com
    
    Automatic get export data of QuestionPro to calculate on Dashboard
"""

# Import package
import os
import openpyxl
import csv
import datetime
import subprocess
from tkinter import *
from tkinter.filedialog import askopenfile

"""
    Object
"""

# Contestant Object
class Contestant:
    # Constructor
    def __init__(self, name):
        self.name = name
        self.judges = []
    
    # Check judge has in list?
    def isJudgeInList(self, judgeName):
        for judge in self.judges:
            if judgeName == judge.name:
                return False
        return True
    
    # Get Result calculate by Judge submission
    def getResult(self):
        # Constant variable
        result = 0
        # variable
        avgJudges = []

        # Get judge not duplicate and have submission
        for judge in self.judges:
            if judge.submission != 0 and judge.duplicate == False:
                avgJudges.append(judge)

        # Check avgJudges have value, if not use current judges list
        if len(avgJudges) == 0:
            avgJudges = self.judges

        for judge in self.judges:
            if judge.duplicate == False:
                result += int(float(judge.submission))
        return float(f"{result / len(avgJudges):.2f}") 
    
# Judge Object
class Judge:
    def __init__(self, name, submission, isSubmission = True):
        self.name = name
        self.submission = submission
        self.duplicate = False
        self.isSubmission = isSubmission

# Dashboard object
class Dashboard:
    def __init__(self):
        self.contestants = []

"""
    Function
"""

# Convert CSV file to Excel file
def convertCSVtoExcel():
    # Create new Excel file
    wb = openpyxl.Workbook()
    # Go to active sheet
    # sheet = wb.active
    sheet = wb["Raw Data"]
    
    # Open CSV file
    with open(file + '.csv') as f:
        reader = csv.reader(f, delimiter=',')
        for row in reader:
            sheet.append(row)

    wb.save(file + '.xlsx')

# Get data by column name
def getDataByColumnName(colName, isConvertToExcel):
    # Variable to return
    result =  []
    index = 0

    # Open QuestionPro file
    wb = openpyxl.load_workbook(file + '.xlsx')
    # Go to active sheet
    if isConvertToExcel:
        sheet = wb.active
    else:
        sheet = wb["Raw Data"]

    # Scan all column
    for column in sheet.iter_cols():
        # Get index of start data
        for i in range(len(column)):
            if "Response ID" in str(column[i]. value):
                index = i
                break
        # Get column name
        columnName = column[index].value
        # Check column name to get data
        if colName in str(columnName):
            # Loop data of column from row 5
            for cell in column[(index + 1):]:
                # Add data to list
                if(colName in 'Hello'):
                    # Check if has value or None then default value
                    result.append(cell.value or 0)
                else:
                    result.append(cell.value or '')
                
    return result

# Convert judge Id to judge name
def convertToJudgeName(judgeId):
    judgesLookup = {
        '0' : '',
        '1' : 'Judge 1 - Invest Barrie',
        '2' : 'Judge 2 - Empower Simcoe',
        '3' : 'Judge 3 - UpChuckle Education',
        '4' : 'Judge 4 - UpLift Black',
        '5' : 'Judge 5 - Georgian College'
    }
    return judgesLookup[str(int(judgeId))]

# Add judge not in data file in Contestant Object judges list with submission 0
def addJudgeNotInData(judges):
    # Constant variables
    totalJudges = ['Judge 1 - Invest Barrie', 'Judge 2 - Empower Simcoe', 'Judge 3 - UpChuckle Education', 'Judge 4 - UpLift Black', 'Judge 5 - Georgian College']
    # Variable
    currentJudges = []

    # Loop judges is having
    for judge in judges:
        # Add judge name having to variable
        currentJudges.append(judge.name)

    # Loop total judges
    for judge in totalJudges:
        # Check if judge not in current judge is having
        if not judge in currentJudges:
            # Add judge not having with submission 0
            judges.append(Judge(judge, 0, False))

# Generate, design and export new Dashboard file
def exportDashboardFile(data):
    # Constant variables
    columnLetters = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']
    
    # Get current datetime
    currentDatetime = datetime.datetime.today().strftime('%Y-%m-%d-%H.%M.%S')

    # Create new Excel file
    dashboardWb = openpyxl.Workbook()
    # Go to active sheet
    dashboard = dashboardWb.active

    # Add default content on cell
    dashboard['A1'] = 'Further Faster Grand Pitch Event'
    dashboard['A3'] = 'Contestant Name'
    dashboard['B3'] = 'Result'
    dashboard['D' + str(len(getJudgesListUniqueSortASC(data)) + 4)] = 'Result'
    dashboard['D3'] = 'Metrix'
    dashboard[columnLetters[len(data.contestants) + 5] + '3'] = 'Duplicate'

    # Add data value on file
    generateRankingData(dashboard, columnLetters, data)
    generateContestantData(dashboard, columnLetters, data)
    generateJudgeData(dashboard, data)
    generateSubmissionData(dashboard, columnLetters, data)
    generateResultData(dashboard, columnLetters, data)
    generateDuplicateData(dashboard, columnLetters, data)

    # Design layout on file
    designLayoutTitle(dashboard, columnLetters, data)

    # Save Dashboard file with current date
    dashboardWb.save(directory + 'Dashboard_' + currentDatetime + '.xlsx')

    # Auto Open Dashboard file
    os.startfile(directory + 'Dashboard_' + currentDatetime + '.xlsx')

# Return judges list with unique and sort ASC list
def getJudgesListUniqueSortASC(data):
    # Variable
    judges = []

    # Loop all contestant from data
    for contestant in data.contestants:
        # Loop all judge from contestant
        for judge in contestant.judges:
            # Add judge name to variable
            judges.append(judge.name)

    # return new judge list with unique and sort ASC
    return sorted(set(judges))

# Add Contestant and Result data to Dashboard file on First Table
def generateRankingData(dashboard, colLetters, data):
    # Variable
    ranks = []

    # Loop all contestant from data
    for contestant in data.contestants:
        # Add contestant name and result on ranking variable
        ranks.append([contestant.name, contestant.getResult()])

    # Raking contestant by result by sort ASC
    for i in range(len(ranks)):
        for j in range(i, len(ranks)):
            if ranks[i][1] < ranks[j][1]:
                temp = ranks[i]
                ranks[i] = ranks[j]
                ranks[j] = temp

    # Loop all ranks have sort
    for i in range(len(ranks)):
        # Loop all item in rank
        for r in range(len(ranks[i])):
            # Add contestant and result to cell from A4 B4 vertical
            dashboard[colLetters[r] + str(i + 4)] = ranks[i][r]
            # Design layout table
            designBorderWidthFirstTable(dashboard, colLetters, r, i)

# Add Contestant data to Dashboard file
def generateContestantData(dashboard, colLetters, data):
    # Loop all contestant from data
    for i in range(len(data.contestants)):
        # Add contestant to cell from E3 horizon
        dashboard[colLetters[i + 4] + '3'] = data.contestants[i].name
        # Design layout
        designBorderWidthContestantsSecondTable(dashboard, colLetters, i)

# Add Judge data to Dashboard file
def generateJudgeData(dashboard, data):
    # Get judge list with unique and sort ASC
    judges = getJudgesListUniqueSortASC(data)

    # Loop all judge from data
    for i in range(len(judges)):
        # Add judge to cell from D4 vertical
        dashboard['D' + str(i + 4)] = judges[i]
        # Design layout
        designBorderWidthJudgesSecondTable(dashboard, i)

# Add Submission data to Dashboard file
def generateSubmissionData(dashboard, colLetters, data):
    # Get judge list with unique and sort ASC
    judges = getJudgesListUniqueSortASC(data)

    # Loop all contestant from data
    for c in range(len(data.contestants)):
        # Loop all judge from judge list unique and sort
        for i in range(len(judges)):
            # Loop all judge from contestant
            for judge in data.contestants[c].judges:
                # Check judge name like judgeUnique and not duplicate
                if judge.name == judges[i] and judge.duplicate == False:
                    # Add submission to cell from E4 vertical and horizon
                    dashboard[colLetters[c + 4] + str(i + 4)] = judge.submission
                    # Design layout
                    designBorderWidthSubmissionsSecondTable(dashboard, colLetters, judge, c, i)

# Add Result data to Dashboard file on Second Table
def generateResultData(dashboard, colLetters, data):
    judgesLength = len(getJudgesListUniqueSortASC(data)) + 4
    # Loop all contestant from data
    for c in range(len(data.contestants)):
        # Add Result to cell from E9 horizon
        dashboard[colLetters[c + 4] + str(judgesLength)] = data.contestants[c].getResult()
        # Design layout
        designBorderWidthResultsSecondTable(dashboard, colLetters, c, judgesLength)

# Add Duplicate data to Dashboard file on Third Table
def generateDuplicateData(dashboard, colLetters, data):
    # Variable
    step = 0
    oldStep = 0
    
    # Loop all contestant from data
    for contestant in data.contestants:
        # Loop all judge from data
        for judge in contestant.judges:
            # Check judge is duplicate
                if judge.duplicate == True:
                    # Add judge to cell from I4 vertical
                    dashboard[colLetters[len(data.contestants) + 6] + str(step + 4)] = judge.name
                    # Add submission to cell from J4 vertical
                    dashboard[colLetters[len(data.contestants) + 7] + str(step + 4)] = judge.submission
                    # Design Layout
                    designBorderWidthJudgesSubmissionsThirdTable(dashboard, colLetters, data.contestants, step,)
                    # Add 1 when have judge duplicate
                    step = step + 1
        # Check have judge duplicate
        if step != oldStep:
            # Add contestant to cell from H4 vertical
            dashboard[colLetters[len(data.contestants) + 5] + str(step + 3)] = contestant.name
            # Design layout
            designBorderWidthContestantsThirdTable(dashboard, colLetters, data.contestants, step)
        # Assign current step to oldStep
        oldStep = step

# Design layout Dashboard file
def designLayoutTitle(dashboard, colLetters, data):
    # Variable to define border
    side = openpyxl.styles.Side(border_style="thin", color="000000")

    # Heading center, bold, font size, merge
    dashboard['A1'].alignment = openpyxl.styles.Alignment(horizontal='center')
    dashboard['A1'].font = openpyxl.styles.Font(bold=True, size = '20')
    # Make border on Heading
    for i in range(len(data.contestants) + 8):
        # If in first cell
        if i == 0:
            dashboard['A1'].border = openpyxl.styles.Border(top=side, left=side, bottom=side)
        # If in last cell
        elif i == (len(data.contestants) + 7):
            dashboard[colLetters[i] + '1'].border = openpyxl.styles.Border(top=side, right=side, bottom=side)
        # If in another cell
        else:
            dashboard[colLetters[i] + '1'].border = openpyxl.styles.Border(top=side, bottom=side)
    # Calculate by contestant generate horizon on Second Table
    dashboard.merge_cells('A1:' + colLetters[len(data.contestants) + 7] + '1')
    

    # Make border and width Title on First Table
    dashboard['A3'].border = openpyxl.styles.Border(top=side, left=side, right=side, bottom=side)
    dashboard.column_dimensions['A'].width = 30
    dashboard['B3'].border = openpyxl.styles.Border(top=side, left=side, right=side, bottom=side)
    dashboard.column_dimensions['B'].width = 30

    # Make border and width Title on Second Table
    dashboard['D3'].border = openpyxl.styles.Border(top=side, left=side, right=side, bottom=side)
    dashboard['D' + str(len(getJudgesListUniqueSortASC(data)) + 4)].border = openpyxl.styles.Border(top=side, left=side, right=side, bottom=side)
    dashboard.column_dimensions['D'].width = 30

    # Heading center, merge on Third Table
    dashboard[colLetters[len(data.contestants) + 5] + '3'].alignment = openpyxl.styles.Alignment(horizontal='center')
    # Make border and width on Third Table
    dashboard[colLetters[len(data.contestants) + 5] + '3'].border = openpyxl.styles.Border(top=side, left=side, bottom=side)
    dashboard[colLetters[len(data.contestants) + 6] + '3'].border = openpyxl.styles.Border(top=side, bottom=side)
    dashboard[colLetters[len(data.contestants) + 7] + '3'].border = openpyxl.styles.Border(top=side, right=side, bottom=side)
    dashboard.column_dimensions[colLetters[len(data.contestants) + 5]].width = 30
    dashboard.column_dimensions[colLetters[len(data.contestants) + 6]].width = 30
    dashboard.column_dimensions[colLetters[len(data.contestants) + 7]].width = 30
    # Calculate by contestant generate horizon on Second Table
    dashboard.merge_cells(colLetters[len(data.contestants) + 5] + '3:' + colLetters[len(data.contestants) + 7] + '3')

# Design Border and Width for Fist Table on Dashboard file
def designBorderWidthFirstTable(dashboard, colLetters, r, i):
    # Variable to define border
    side = openpyxl.styles.Side(border_style="thin", color="000000")
    # Variable for color top 3
    colors = ['4fad5b', 'ffff54', 'ea3323']

    # Check i not out of index
    if i > len(colors) - 1:
        rgb = 'ffffff'
    else:
        rgb = colors[i]

    # Make border and width
    dashboard[colLetters[r] + str(i + 4)].border = openpyxl.styles.Border(left=side, right=side, bottom=side)

    # Fill color for top 3
    dashboard[colLetters[r] + str(i + 4)].fill = openpyxl.styles.PatternFill(patternType='solid', 
                                                    fgColor=openpyxl.styles.colors.Color(rgb=rgb))
    
# Design Border and Width for Judges on Second Table on Dashboard file
def designBorderWidthJudgesSecondTable(dashboard, i):
    # Variable to define border
    side = openpyxl.styles.Side(border_style="thin", color="000000")

    # Make border and width
    dashboard['D' + str(i + 4)].border = openpyxl.styles.Border(left=side, right=side)

# Design Border and Width for Contestants on Second Table on Dashboard file
def designBorderWidthContestantsSecondTable(dashboard, colLetters, i):
    # Variable to define border
    side = openpyxl.styles.Side(border_style="thin", color="000000")

    # Make border and width
    dashboard[colLetters[i + 4] + '3'].border = openpyxl.styles.Border(left=side, right=side, bottom=side, top=side)
    dashboard.column_dimensions[colLetters[i + 4]].width = 30

# Design Border and Width for Submissions on Second Table on Dashboard file
def designBorderWidthSubmissionsSecondTable(dashboard, colLetters, judge, c, i):
    # Variable to define border
    side = openpyxl.styles.Side(border_style="thin", color="000000")

    # Make border and width
    dashboard[colLetters[c + 4] + str(i + 4)].border = openpyxl.styles.Border(left=side, right=side)

    # Fill red color on data not have submission
    if judge.isSubmission == False:
        dashboard[colLetters[c + 4] + str(i + 4)].fill = openpyxl.styles.PatternFill(patternType='solid', 
                                                    fgColor=openpyxl.styles.colors.Color(rgb='ea3323'))

# Design Border and Width for Results on Second Table on Dashboard file
def designBorderWidthResultsSecondTable(dashboard, colLetters, c, judgesLength):
    # Variable to define border
    side = openpyxl.styles.Side(border_style="thin", color="000000")

    # Make border and width
    dashboard[colLetters[c + 4] + str(judgesLength)].border = openpyxl.styles.Border(left=side, right=side, bottom=side, top=side)

# Design Border and Width for Judges, Submissions on Third Table on Dashboard file
def designBorderWidthJudgesSubmissionsThirdTable(dashboard, colLetters, contestant, step):
    # Variable to define border
    side = openpyxl.styles.Side(border_style="thin", color="000000")

    # Make border and width
    dashboard[colLetters[len(contestant) + 6] + str(step + 4)].border = openpyxl.styles.Border(bottom=side, top=side)
    dashboard[colLetters[len(contestant) + 7] + str(step + 4)].border = openpyxl.styles.Border(bottom=side, top=side, right=side)

# Design Border and Width for Contestants on Third Table on Dashboard file
def designBorderWidthContestantsThirdTable(dashboard, colLetters, contestant, step):
    # Variable to define border
    side = openpyxl.styles.Side(border_style="thin", color="000000")

    # Make border and width
    dashboard[colLetters[len(contestant) + 5] + str(step + 3)].border = openpyxl.styles.Border(left=side, right=side, bottom=side, top=side)

def openFile():
    # Get global file
    global file, directory

    # Open File Explore to choose file
    fileSelect = askopenfile(title = "Select file", filetypes =[("CSV Files", "*.csv"),("Excel files", ".xlsx")])
    # Check file choosen
    if fileSelect:
        # Assign file with no file type
        file = os.path.basename(fileSelect.name)
        # Assign directory file located
        directory = os.path.dirname(fileSelect.name) + "/"
        # Show File Name on Label
        labelFilename["text"] = os.path.basename(fileSelect.name)

# Function button Calculate Click
def calculateFile():
    # Get global file
    global file, directory

    # Constant variables
    contestantColName = 'Custom Variable 2'
    judgeColName = 'Hello'
    submissionColName = 'Weight'
    statusColName = 'Response Status'
    isConvertToExcel = False

    # Data variable
    dashboard = Dashboard()

    # Check user has choose file
    if file:
        if ".csv" in file:
            file = file.rstrip(".csv")
            # Assign file user choose
            file = directory + file 

            # Convert CSV to Excel file
            convertCSVtoExcel()
            isConvertToExcel = True
        else:
            file = file.rstrip(".xlsx")
            # Assign file user choose
            file = directory + file

        # Get data
        contestants = getDataByColumnName(contestantColName, isConvertToExcel)
        judges = getDataByColumnName(judgeColName, isConvertToExcel)
        submissions = getDataByColumnName(submissionColName, isConvertToExcel)
        status = getDataByColumnName(statusColName, isConvertToExcel)

        # Get contestant in list of unique and sort ASC from contestants
        for contestantUnique in sorted(set(contestants)):
            # Create new contestant object
            contestant = Contestant(contestantUnique)
            # Loop contestants data from end to start
            for i in range(len(contestants) - 1, -1, -1):
                # If the same contestants
                if contestantUnique == contestants[i] and str(status[i]) in 'Completed':
                    # Create new judge object
                    judge = Judge(convertToJudgeName(judges[i]), int(float(submissions[i])))
                    # Check judge have in list of contestant object
                    if not contestant.isJudgeInList(judge.name):
                        # Update duplicate field of judge to True
                        judge.duplicate = True
                    # Add judge list to contestant object
                    contestant.judges.append(judge)
        
            # Add judge not having in list
            addJudgeNotInData(contestant.judges)

            # Add contestant to list of dashboard object
            dashboard.contestants.append(contestant)

        # Create, generate and auto open result file
        exportDashboardFile(dashboard)

"""
    Main
"""

# Constant variables
directory = ''
file = ''

# Create app dialog
app = Tk()

# Add title to app
app.title("Further Faster Grand Pitch")
# Set width and height of app
app.geometry("500x200")

# Add Select File button with function click
btnOpenFile = Button(app, text = "Select File", command=openFile)
# Position button on app and padding vertical
btnOpenFile.pack(pady=(20, 20))

# Add Calculate button with function click
btnCalculateFile = Button(app, text = "Calculate File", command=calculateFile)
# Position button on app and padding vertical
btnCalculateFile.pack(pady=(0, 20))

# Add Label File Name to know which File selected
labelFilename = Label(app, text = "")
# Position button on app and full width
labelFilename.pack(fill="x")

# Execute app
app.mainloop()
